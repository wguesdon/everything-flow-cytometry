#!/usr/bin/env python3
"""Derive the committed metadata of FlowRepository FR-FCM-Z282.

The deposit carries two sources of design information. The first is the attached
spreadsheet `De_Identified_Proficiency_naive_mem_DataBase.xlsx`, which holds the
frequency that each operator reported and the frequency that the reference
operator measured from the same file. The second is the FCS file name, which
encodes the operator, the instrument, the centre, the instrument model, the
material and the round.

This script writes three CSV files into `gating/`. They are committed, so the
analysis runs without the 5.8 GB deposit and without a spreadsheet reader.

Run it once, on the host, after `./sync.sh pull`:

    uv run --project python python scripts/z282_derive_metadata.py
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

DEPOSIT = Path(
    "data/datasets/flowrepository/FR-FCM-Z282/FlowRepository_FR-FCM-Z282_files"
)
DATABASE = DEPOSIT / "attachments" / "De_Identified_Proficiency_naive_mem_DataBase.xlsx"
OUTPUT_DIR = Path("gating")

# A sample file name looks like OpA_I2_C2_IM1_PBMC1_R1.fcs. Two operators wrote a
# hyphen in place of the fourth separator, so the pattern accepts both.
SAMPLE_PATTERN = re.compile(
    r"^Op(?P<operator>[A-M])"
    r"_I(?P<instrument>\d+)"
    r"_C(?P<centre>\d+)"
    r"_IM(?P<model>\d+)[_-]"
    r"(?P<material>PBMC|WB)(?P<vial>\d+)"
    r"_R(?P<round>\d+)\.fcs$"
)

# The remaining files are the single stain controls of the reference operator.
CONTROL_PATTERN = re.compile(
    r"^Op(?P<operator>[A-M])"
    r"_I(?P<instrument>\d+)"
    r"_C(?P<centre>\d+)"
    r"_IM(?P<model>\d+)_"
    r"(?P<material>PBMC|WB)(?P<vial>\d+)_"
    r"(?P<stain>.+)\.fcs$"
)

# The fifteen frequencies that the spreadsheet reports for every row. The paper
# counts its acceptability results out of this list, so the order is kept.
PARAMETERS = (
    "CD3", "CD4", "CD8",
    "CD3_N", "CD4_N", "CD8_N",
    "CD3_CM", "CD4_CM", "CD8_CM",
    "CD3_EM", "CD4_EM", "CD8_EM",
    "CD3_TD", "CD4_TD", "CD8_TD",
)


def read_database(path: Path) -> list[dict[str, str]]:
    """Read the proficiency spreadsheet into long form.

    The sheet is wide. Each row carries one observation of fifteen parameters.
    The long form gives one row per parameter, which is what every statistic in
    the report consumes.

    Args:
        path: Path to the attached xlsx file.

    Returns:
        A list of dictionaries, one per operator, material, vial, round,
        analysis type and parameter.

    Raises:
        FileNotFoundError: If the spreadsheet is not in the deposit.
        ValueError: If the header does not carry every expected column.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"The spreadsheet is missing: {path}\n"
            "Run ./sync.sh pull datasets/flowrepository/FR-FCM-Z282 first."
        )
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value) for value in rows[0]]
    missing = [name for name in PARAMETERS if name not in header]
    if missing:
        raise ValueError(f"The spreadsheet header lacks these columns: {missing}")

    index = {name: header.index(name) for name in header}
    long_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        if row[index["Random_Op_id"]] is None:
            continue
        for parameter in PARAMETERS:
            value = row[index[parameter]]
            long_rows.append(
                {
                    "operator": str(row[index["Random_Op_id"]]),
                    "centre": str(row[index["Centre_no"]]),
                    "instrument": str(row[index["Instrument_no"]]),
                    "instrument_model": str(row[index["Instrument_Model"]]),
                    "analysis": str(row[index["Analysis_Type"]]),
                    "material": str(row[index["PBMC_WB"]]),
                    "vial": str(row[index["Vial_no"]]),
                    "round": str(row[index["Round_no"]]),
                    "excluded": str(row[index["Exclusion"]]),
                    "parameter": parameter,
                    "percent": "" if value is None else str(value),
                }
            )
    return long_rows


def read_file_names(deposit: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Parse the design out of every FCS file name in the deposit.

    Args:
        deposit: Path to the extracted `FlowRepository_FR-FCM-Z282_files` folder.

    Returns:
        A pair. The first item is one row per sample file. The second item is one
        row per single stain control file.

    Raises:
        FileNotFoundError: If the deposit folder is not present.
        ValueError: If a file name matches neither pattern.
    """
    if not deposit.is_dir():
        raise FileNotFoundError(
            f"The deposit is missing: {deposit}\n"
            "Run ./sync.sh pull datasets/flowrepository/FR-FCM-Z282 first."
        )
    samples: list[dict[str, str]] = []
    controls: list[dict[str, str]] = []
    for path in sorted(deposit.glob("*.fcs")):
        match = SAMPLE_PATTERN.match(path.name)
        if match is not None:
            fields = match.groupdict()
            samples.append({"file_name": path.name, **fields})
            continue
        match = CONTROL_PATTERN.match(path.name)
        if match is None:
            raise ValueError(f"This file name matches no pattern: {path.name}")
        fields = match.groupdict()
        controls.append({"file_name": path.name, **fields})
    return samples, controls


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    """Write a list of dictionaries to a CSV file.

    Args:
        path: Destination path.
        rows: The rows to write. The first row sets the column order.

    Raises:
        ValueError: If the row list is empty.
    """
    if not rows:
        raise ValueError(f"There is nothing to write to {path}")
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    """Derive the three metadata files and report what was written.

    Returns:
        0 when every file was written.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--deposit", type=Path, default=DEPOSIT, help="The extracted deposit folder"
    )
    parser.add_argument(
        "--output", type=Path, default=OUTPUT_DIR, help="Where to write the CSV files"
    )
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)

    results = read_database(args.deposit / "attachments" / DATABASE.name)
    samples, controls = read_file_names(args.deposit)

    write_csv(args.output / "z282_operator_results.csv", results)
    write_csv(args.output / "z282_sample_metadata.csv", samples)
    write_csv(args.output / "z282_control_files.csv", controls)

    print(f"z282_operator_results.csv  {len(results)} rows")
    print(f"z282_sample_metadata.csv   {len(samples)} sample files")
    print(f"z282_control_files.csv     {len(controls)} single stain controls")
    return 0


if __name__ == "__main__":
    sys.exit(main())
